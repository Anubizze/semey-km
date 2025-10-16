from flask import Flask, render_template, request, redirect, url_for, session, jsonify, make_response
from models import db, User, Entry, Batch, Equipment, Material, BatchMaterial, Product, BatchTemplate, BatchTemplateMaterial
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import csv
import io
import json
import os
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = "supersecretkey"
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///factory.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db.init_app(app)

# === СИСТЕМА КЭШИРОВАНИЯ ===
class SimpleCache:
    def __init__(self):
        self.cache = {}
        self.cache_ttl = {}  # Time To Live
    
    def get(self, key):
        if key in self.cache:
            if time.time() < self.cache_ttl.get(key, 0):
                return self.cache[key]
            else:
                # Кэш истек
                del self.cache[key]
                del self.cache_ttl[key]
        return None
    
    def set(self, key, value, ttl_seconds=300):  # 5 минут по умолчанию
        self.cache[key] = value
        self.cache_ttl[key] = time.time() + ttl_seconds
    
    def clear(self):
        self.cache.clear()
        self.cache_ttl.clear()
    
    def clear_pattern(self, pattern):
        """Очищает кэш по паттерну ключа"""
        keys_to_remove = [key for key in self.cache.keys() if pattern in key]
        for key in keys_to_remove:
            del self.cache[key]
            del self.cache_ttl[key]

# Глобальный кэш
cache = SimpleCache()

def get_cached_data(key, fetch_func, ttl_seconds=300):
    """Получает данные из кэша или выполняет функцию и кэширует результат"""
    cached_data = cache.get(key)
    if cached_data is not None:
        return cached_data
    
    # Выполняем функцию и кэшируем результат
    data = fetch_func()
    cache.set(key, data, ttl_seconds)
    return data

def clear_analytics_cache():
    """Очищает кэш аналитических данных"""
    cache.clear_pattern("material_consumption")
    cache.clear_pattern("daily_analytics")
    cache.clear_pattern("weekly_analytics")

@app.route("/")
def index():
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        user = User.query.filter_by(username=username).first()
        
        # Отладочная информация
        print(f"Попытка входа: {username}")
        if user:
            print(f"Пользователь найден: {user.username}, роль: {user.role}")
            print(f"Пароль в базе: {user.password}")
            print(f"Введенный пароль: {password}")
            
            # Проверяем пароль (поддерживаем как хешированные, так и простые пароли)
            if check_password_hash(user.password, password) or user.password == password:
                print("Пароль верный!")
            session["user_id"] = user.id
            session["role"] = user.role

            if user.role == "admin":
                return redirect(url_for("admin_dashboard"))
            elif user.role in ["director", "chief_technologist"]:
                return redirect(url_for("director_dashboard"))
            elif user.role == "autoclave_operator":
                return redirect(url_for("autoclave_dashboard"))
            elif user.role == "cutting_operator":
                return redirect(url_for("cutting_dashboard"))
            elif user.role == "casting_operator":
                return redirect(url_for("employee_dashboard"))  # Заливщик вводит материалы
            else:
                # Обычный сотрудник
                return redirect(url_for("employee_dashboard"))
        else:
            print("Пароль неверный!")
            return render_template("login.html", error="Неверный пароль")
    else:
        print("Пользователь не найден!")
        return render_template("login.html", error="Пользователь не найден")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/admin_dashboard")
def admin_dashboard():
    if session.get("role") != "admin":
        return redirect(url_for("login"))

    employees = User.query.all()
    return render_template("admin_dashboard.html", employees=employees)

@app.route("/add_employee", methods=["GET", "POST"])
def add_employee():
    if session.get("role") != "admin":
        return redirect(url_for("login"))

    if request.method == "POST":
        fio = request.form["fio"]
        iin = request.form["iin"]
        position = request.form["position"]
        role = request.form["role"]

        username = iin
        raw_password = iin[-6:]
        password = generate_password_hash(raw_password)

        if User.query.filter_by(username=username).first():
            return "Пользователь с таким ИИН уже существует", 400

        new_user = User(fio=fio, iin=iin, username=username, password=password, role=role, position=position)
        db.session.add(new_user)
        db.session.commit()

        return redirect(url_for("admin_dashboard"))

    return render_template("add_employee.html")

@app.route("/delete_employee/<int:id>", methods=["POST"])
def delete_employee(id):
    if session.get("role") != "admin":
        return redirect(url_for("login"))

    employee = User.query.get_or_404(id)
    db.session.delete(employee)
    db.session.commit()

    return redirect(url_for("admin_dashboard"))

# Управление продуктами
@app.route("/products")
def products_list():
    if session.get("role") != "admin":
        return redirect(url_for("login"))
    
    products = Product.query.filter_by(is_active=True).all()
    return render_template("products_list.html", products=products)

@app.route("/add_product", methods=["GET", "POST"])
def add_product():
    if session.get("role") != "admin":
        return redirect(url_for("login"))
    
    if request.method == "POST":
        product = Product(
            name=request.form["name"],
            product_code=request.form["product_code"],
            description=request.form.get("description", "")
        )
        
        try:
            db.session.add(product)
            db.session.commit()
            return redirect(url_for("products_list"))
        except Exception as e:
            db.session.rollback()
            return f"Ошибка: {str(e)}", 400
    
    return render_template("add_product.html")

@app.route("/delete_product/<int:id>", methods=["POST"])
def delete_product(id):
    if session.get("role") != "admin":
        return redirect(url_for("login"))
    
    product = Product.query.get_or_404(id)
    product.is_active = False  # Мягкое удаление
    db.session.commit()
    
    return redirect(url_for("products_list"))

@app.route("/employee_dashboard", methods=["GET", "POST"])
def employee_dashboard():
    allowed_roles = ["employee", "casting_operator"]
    if "user_id" not in session or session.get("role") not in allowed_roles:
        return redirect(url_for("login"))

    if request.method == "POST":
        entry = Entry(
            user_id=session["user_id"],
            cement=request.form.get("cement", type=float, default=0),
            lime=request.form.get("lime", type=float, default=0),
            alum_powder=request.form.get("alumina_powder", type=float, default=0),
            sludge=request.form.get("sludge", type=float, default=0),
            gypsum=request.form.get("gypsum", type=float, default=0),
            water=request.form.get("water", type=float, default=0),
            sulfanol=request.form.get("sulfanol", type=float, default=0),
            time=datetime.now().strftime("%H:%M:%S"),
            shift=request.form.get("shift", "day"),
            date=datetime.now().date()
        )
        db.session.add(entry)
        db.session.commit()
        clear_analytics_cache()  # Очищаем кэш при добавлении новых данных

        return redirect(url_for("employee_dashboard"))

    # Получаем историю операций за сегодня
    today = datetime.now().date()
    today_entries = Entry.query.filter_by(
        user_id=session["user_id"], 
        date=today
    ).order_by(Entry.time.desc()).limit(10).all()
    
    return render_template("employee_dashboard.html", today_entries=today_entries)

# Дашборд для резчика (cutting_operator)
@app.route("/cutting_dashboard", methods=["GET", "POST"])
def cutting_dashboard():
    if "user_id" not in session or session.get("role") != "cutting_operator":
        return redirect(url_for("login"))
    
    # Получаем активные партии резки этого пользователя
    active_batches = Batch.query.filter_by(
        user_id=session["user_id"],
        batch_type="cutting",
        status="active"
    ).all()
    
    # Получаем завершенные партии
    completed_batches = Batch.query.filter_by(
        user_id=session["user_id"],
        batch_type="cutting",
        status="completed"
    ).order_by(Batch.end_time.desc()).limit(10).all()
    
    # Получаем доступное оборудование (резчики)
    cutters = Equipment.query.filter_by(
        equipment_type="cutting",
        status="operational"
    ).all()
    
    # Получаем историю операций резки за сегодня
    today = datetime.now().date()
    today_batches = Batch.query.filter_by(
        user_id=session["user_id"],
        batch_type="cutting"
    ).filter(
        db.func.date(Batch.start_time) == today
    ).order_by(Batch.start_time.desc()).limit(10).all()
    
    return render_template("cutting_dashboard.html",
                         active_batches=active_batches,
                         completed_batches=completed_batches,
                         cutters=cutters,
                         today_batches=today_batches)

# Дашборд для автоклавщика (autoclave_operator)
@app.route("/autoclave_dashboard", methods=["GET", "POST"])
def autoclave_dashboard():
    if "user_id" not in session or session.get("role") != "autoclave_operator":
        return redirect(url_for("login"))
    
    # Получаем активные циклы автоклавирования
    active_batches = Batch.query.filter_by(
        user_id=session["user_id"],
        batch_type="autoclave",
        status="active"
    ).all()
    
    # Получаем завершенные циклы
    completed_batches = Batch.query.filter_by(
        user_id=session["user_id"],
        batch_type="autoclave",
        status="completed"
    ).order_by(Batch.end_time.desc()).limit(10).all()
    
    # Получаем доступное оборудование (автоклавы)
    autoclaves = Equipment.query.filter_by(
        equipment_type="autoclave",
        status="operational"
    ).all()
    
    return render_template("autoclave_dashboard.html",
                         active_batches=active_batches,
                         completed_batches=completed_batches,
                         autoclaves=autoclaves)

@app.route("/director_dashboard")
def director_dashboard():
    allowed_roles = ["director", "chief_technologist"]
    if session.get("role") not in allowed_roles:
        return redirect(url_for("login"))

    # Получаем параметры фильтрации
    from datetime import datetime, timedelta
    
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    batch_type_filter = request.args.get('batch_type', 'all')
    status_filter = request.args.get('status', 'all')
    product_filter = request.args.get('product', 'all')
    equipment_filter = request.args.get('equipment', 'all')
    
    # Устанавливаем значения по умолчанию (последние 7 дней)
    if not date_from:
        date_from = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = datetime.now().strftime('%Y-%m-%d')

    # Получаем все записи с данными пользователей (с фильтром по датам)
    entries_query = db.session.query(Entry, User).join(User, Entry.user_id == User.id)
    
    # Применяем фильтр по датам для Entry
    try:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        entries_query = entries_query.filter(Entry.date >= date_from_obj, Entry.date <= date_to_obj)
    except ValueError:
        pass  # Если дата некорректна, игнорируем фильтр
    
    entries = entries_query.all()
    
    # Базовый запрос для партий с фильтром по датам
    batches_query = db.session.query(Batch, User).join(User, Batch.user_id == User.id)
    
    # Применяем фильтр по датам для Batch
    try:
        batches_query = batches_query.filter(
            Batch.start_time >= datetime.strptime(date_from, '%Y-%m-%d'),
            Batch.start_time <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
        )
    except ValueError:
        pass
    
    # Статистика по сменам (Entry записи - заливка)
    day_entries = [e for e, u in entries if e.shift == "day"]
    night_entries = [e for e, u in entries if e.shift == "night"]
    
    # Статистика по сменам для Batch записей (резка и автоклав)
    all_batches_for_shifts = batches_query.all()
    day_batches = [b for b, u in all_batches_for_shifts if b.shift == "day"]
    night_batches = [b for b, u in all_batches_for_shifts if b.shift == "night"]
    
    # Общая статистика по сменам (Entry + Batch)
    total_day_operations = len(day_entries) + len(day_batches)
    total_night_operations = len(night_entries) + len(night_batches)
    
    # Общая статистика по материалам
    total_entries = len(entries)
    total_cement = sum(e.cement for e, u in entries)
    total_lime = sum(e.lime for e, u in entries)
    total_water = sum(e.water for e, u in entries)
    
    # Применяем фильтр по типу партии
    if batch_type_filter != 'all':
        batches_query_filtered = batches_query.filter(Batch.batch_type == batch_type_filter)
    else:
        batches_query_filtered = batches_query
    
    # Применяем фильтр по статусу
    if status_filter != 'all':
        batches_query_filtered = batches_query_filtered.filter(Batch.status == status_filter)
    
    # Применяем фильтр по продукту
    if product_filter != 'all':
        batches_query_filtered = batches_query_filtered.filter(Batch.product_id == product_filter)
    
    # Применяем фильтр по оборудованию
    if equipment_filter != 'all':
        batches_query_filtered = batches_query_filtered.filter(Batch.equipment_id == equipment_filter)
    
    # Аналитика по партиям резки
    cutting_batches_query = batches_query.filter(Batch.batch_type == "cutting")
    cutting_batches = cutting_batches_query.order_by(Batch.start_time.desc()).all()
    total_cutting_batches = len(cutting_batches)
    active_cutting = len([b for b, u in cutting_batches if b.status in ["active", "inactive"]])
    completed_cutting = len([b for b, u in cutting_batches if b.status == "completed"])
    
    # Средняя длительность резки (в минутах)
    cutting_durations = []
    for batch, user in cutting_batches:
        if batch.end_time and batch.start_time:
            duration = (batch.end_time - batch.start_time).total_seconds() / 60
            cutting_durations.append(duration)
    avg_cutting_duration = sum(cutting_durations) / len(cutting_durations) if cutting_durations else 0
    
    # Аналитика по циклам автоклавирования
    autoclave_batches_query = batches_query.filter(Batch.batch_type == "autoclave")
    autoclave_batches = autoclave_batches_query.order_by(Batch.start_time.desc()).all()
    total_autoclave_batches = len(autoclave_batches)
    active_autoclave = len([b for b, u in autoclave_batches if b.status in ["active", "inactive"]])
    completed_autoclave = len([b for b, u in autoclave_batches if b.status == "completed"])
    
    # Средняя длительность автоклавирования (в минутах)
    autoclave_durations = []
    for batch, user in autoclave_batches:
        if batch.end_time and batch.start_time:
            duration = (batch.end_time - batch.start_time).total_seconds() / 60
            autoclave_durations.append(duration)
    avg_autoclave_duration = sum(autoclave_durations) / len(autoclave_durations) if autoclave_durations else 0
    
    # Аналитика по заливке
    casting_batches_query = batches_query.filter(Batch.batch_type == "casting")
    casting_batches = casting_batches_query.all()
    total_casting_batches = len(casting_batches)
    
    # Аналитика по продуктам
    products_stats = {}
    all_batches = batches_query.all()
    for batch, user in all_batches:
        if batch.product:
            product_code = batch.product.product_code
            if product_code not in products_stats:
                products_stats[product_code] = {
                    'name': batch.product.name,
                    'total': 0,
                    'active': 0,
                    'completed': 0
                }
            products_stats[product_code]['total'] += 1
            if batch.status == 'active':
                products_stats[product_code]['active'] += 1
            elif batch.status == 'completed':
                products_stats[product_code]['completed'] += 1
    
    # === ОБЩАЯ АНАЛИТИКА: РАСХОД СЫРЬЯ (СТАТИЧНАЯ) ===
    
    # Используем кэшированные данные о расходе материалов
    material_data = get_cached_data(
        "material_consumption", 
        get_material_consumption_data, 
        ttl_seconds=600  # 10 минут
    )
    
    materials_by_batch_type = material_data["materials_by_batch_type"]
    materials_total = material_data["materials_total"]
    
    # === ЛЕНТА ОПЕРАЦИЙ ЗА ПЕРИОД ===
    
    # Получаем все Entry записи за период
    entries_timeline = []
    if date_from and date_to:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            to_date = datetime.strptime(date_to, '%Y-%m-%d').replace(hour=23, minute=59, second=59)
            
            # Entry записи (ввод данных по заливке)
            entry_records = db.session.query(Entry, User).join(
                User, Entry.user_id == User.id
            ).filter(
                Entry.date >= from_date.date(),
                Entry.date <= to_date.date()
            ).order_by(Entry.date.desc(), Entry.time.desc()).all()
            
            for entry, user in entry_records:
                entries_timeline.append({
                    'type': 'entry',
                    'datetime': datetime.combine(entry.date, datetime.strptime(entry.time, '%H:%M:%S').time()),
                    'user': user,
                    'data': entry,
                    'description': f"Ввод данных заливки - {entry.shift} смена"
                })
            
            # Batch записи (создание партий)
            batch_records = db.session.query(Batch, User).join(
                User, Batch.user_id == User.id
            ).filter(
                Batch.start_time >= from_date,
                Batch.start_time <= to_date
            ).order_by(Batch.start_time.desc()).all()
            
            for batch, user in batch_records:
                status_text = ""
                if batch.status == 'cancelled':
                    status_text = " (ОТМЕНЕНА)"
                elif batch.status == 'completed':
                    status_text = " (ЗАВЕРШЕНА)"
                elif batch.status == 'inactive':
                    status_text = " (ПРИОСТАНОВЛЕНА)"
                
                entries_timeline.append({
                    'type': 'batch',
                    'datetime': batch.start_time,
                    'user': user,
                    'data': batch,
                    'description': f"Создание партии {batch.batch_type} - {batch.batch_number}{status_text}"
                })
            
        except ValueError:
            pass
    
    # Сортируем по времени (новые сначала)
    entries_timeline.sort(key=lambda x: x['datetime'], reverse=True)
    
    # Статистика активности операторов
    operator_stats = {}
    for item in entries_timeline:
        user_id = item['user'].id
        user_name = item['user'].fio
        if user_id not in operator_stats:
            operator_stats[user_id] = {
                'name': user_name,
                'role': item['user'].role,
                'entries_count': 0,
                'batches_count': 0,
                'last_activity': item['datetime']
            }
        
        if item['type'] == 'entry':
            operator_stats[user_id]['entries_count'] += 1
        elif item['type'] == 'batch':
            operator_stats[user_id]['batches_count'] += 1
        
        # Обновляем последнюю активность
        if item['datetime'] > operator_stats[user_id]['last_activity']:
            operator_stats[user_id]['last_activity'] = item['datetime']
    
    # === ПРОИЗВОДСТВЕННЫЕ МЕТРИКИ ===
    
    # 1. Общая выработка (количество партий)
    total_production = len(all_batches)
    
    # 2. Простой (неактивные партии)
    downtime_batches = len([b for b, u in all_batches if b.status == 'inactive'])
    
    # 3. Время работы (активные партии)
    active_time_batches = len([b for b, u in all_batches if b.status == 'active'])
    
    # 4. Количество брака (отмененные партии)
    defect_batches = len([b for b, u in all_batches if b.status == 'cancelled'])
    
    # 5. Нарушения технологий (партии без продукта или оборудования)
    tech_violations = len([b for b, u in all_batches if not b.product_id or not b.equipment_id])
    
    # 6. Эффективность (отношение завершенных к общему количеству)
    completed_batches = len([b for b, u in all_batches if b.status == 'completed'])
    efficiency = (completed_batches / total_production * 100) if total_production > 0 else 0
    
    # 7. Средняя длительность операций
    total_duration = 0
    duration_count = 0
    for batch, user in all_batches:
        if batch.end_time and batch.start_time:
            duration = (batch.end_time - batch.start_time).total_seconds() / 60
            total_duration += duration
            duration_count += 1
    avg_duration = total_duration / duration_count if duration_count > 0 else 0
    
    # === ОБЪЕДИНЕННЫЕ ЗАПИСИ МАТЕРИАЛОВ ===
    
    # Создаем объединенный список записей материалов из всех источников
    material_entries = []
    
    # 1. Записи Entry (заливка)
    for entry, user in entries:
        material_entries.append({
            'type': 'entry',
            'datetime': datetime.combine(entry.date, datetime.strptime(entry.time, '%H:%M:%S').time()),
            'user': user,
            'shift': entry.shift,
            'cement': entry.cement,
            'lime': entry.lime,
            'aluminum_powder': entry.alum_powder,
            'water': entry.water,
            'sulfanol': entry.sulfanol,
            'gypsum': entry.gypsum,
            'sludge': entry.sludge
        })
    
    # 2. Записи Batch с материалами (резка и автоклав)
    for batch, user in all_batches:
        if batch.materials:  # Если у партии есть материалы
            # Получаем материалы для этой партии
            batch_materials = {}
            for bm in batch.materials:
                batch_materials[bm.material.name] = bm.quantity
            
            material_entries.append({
                'type': 'batch',
                'datetime': batch.start_time,
                'user': user,
                'shift': batch.shift,
                'cement': batch_materials.get('Цемент', 0),
                'lime': batch_materials.get('Известь', 0),
                'aluminum_powder': batch_materials.get('Алюминиевая пудра', 0),
                'water': batch_materials.get('Вода', 0),
                'sulfanol': batch_materials.get('Сульфанол', 0),
                'gypsum': batch_materials.get('Гипс', 0),
                'sludge': batch_materials.get('Шлам', 0),
                'batch_type': batch.batch_type,
                'batch_number': batch.batch_number
            })
    
    # Сортируем по времени (новые сначала)
    material_entries.sort(key=lambda x: x['datetime'], reverse=True)
    
    return render_template("director_dashboard.html", 
                         entries=entries,
                         material_entries=material_entries,
                         total_entries=total_entries,
                         total_cement=total_cement,
                         total_lime=total_lime,
                         total_water=total_water,
                         day_entries=total_day_operations,
                         night_entries=total_night_operations,
                         # Детализация по сменам для каждого типа операций
                         day_entries_casting=len(day_entries),
                         night_entries_casting=len(night_entries),
                         day_entries_cutting=len([b for b in day_batches if b.batch_type == "cutting"]),
                         night_entries_cutting=len([b for b in night_batches if b.batch_type == "cutting"]),
                         day_entries_autoclave=len([b for b in day_batches if b.batch_type == "autoclave"]),
                         night_entries_autoclave=len([b for b in night_batches if b.batch_type == "autoclave"]),
                         # Аналитика партий
                         total_cutting_batches=total_cutting_batches,
                         active_cutting=active_cutting,
                         completed_cutting=completed_cutting,
                         avg_cutting_duration=avg_cutting_duration,
                         total_autoclave_batches=total_autoclave_batches,
                         active_autoclave=active_autoclave,
                         completed_autoclave=completed_autoclave,
                         avg_autoclave_duration=avg_autoclave_duration,
                         total_casting_batches=total_casting_batches,
                         cutting_batches=cutting_batches[:10],  # Последние 10
                         autoclave_batches=autoclave_batches[:10],  # Последние 10
                         products_stats=products_stats,  # Статистика по продуктам
                         # НОВОЕ: Расход сырья
                         materials_by_batch_type=materials_by_batch_type,  # Расход по типам партий
                         materials_total=materials_total,  # Общий расход всех материалов
                         # Параметры фильтров
                         date_from=date_from,
                         date_to=date_to,
                         batch_type_filter=batch_type_filter,
                         status_filter=status_filter,
                         product_filter=product_filter,
                         equipment_filter=equipment_filter,
                         # Данные для фильтров
                         products=Product.query.filter_by(is_active=True).all(),
                         equipment=Equipment.query.filter_by(status="operational").all(),
                         # Лента операций
                         entries_timeline=entries_timeline,
                         operator_stats=operator_stats,
                         # Производственные метрики
                         total_production=total_production,
                         downtime_batches=downtime_batches,
                         active_time_batches=active_time_batches,
                         defect_batches=defect_batches,
                         tech_violations=tech_violations,
                         efficiency=efficiency,
                         avg_duration=avg_duration)

@app.route("/analytics_data")
def analytics_data():
    if session.get("role") != "director":
        return redirect(url_for("login"))

    # Получаем данные для графиков
    entries = db.session.query(Entry, User).join(User, Entry.user_id == User.id).all()
    
    # Группируем по датам
    daily_data = {}
    for entry, user in entries:
        date_str = entry.date.strftime("%Y-%m-%d")
        if date_str not in daily_data:
            daily_data[date_str] = {
                'cement': 0, 'lime': 0, 'water': 0, 'count': 0
            }
        daily_data[date_str]['cement'] += entry.cement
        daily_data[date_str]['lime'] += entry.lime
        daily_data[date_str]['water'] += entry.water
        daily_data[date_str]['count'] += 1
    
    return jsonify({
        'daily_data': daily_data,
        'total_entries': len(entries)
    })

# Новые маршруты для системы партий

@app.route("/create_batch", methods=["GET", "POST"])
def create_batch():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    if request.method == "POST":
        batch_number = request.form["batch_number"]
        batch_type = request.form["batch_type"]
        product_id = request.form.get("product_id", type=int)
        equipment_id = request.form.get("equipment_id", type=int)
        notes = request.form.get("notes", "")
        
        # Получаем время начала и окончания (если указано)
        start_time_str = request.form.get("start_time", "")
        end_time_str = request.form.get("end_time", "")
        
        # Валидация времени
        start_time = datetime.now()
        end_time = None
        
        if start_time_str:
            try:
                start_time = datetime.strptime(start_time_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                return "Неверный формат времени начала", 400
        
        if end_time_str:
            try:
                end_time = datetime.strptime(end_time_str, '%Y-%m-%dT%H:%M')
                # Проверяем, что время окончания больше времени начала
                if end_time <= start_time:
                    return "Время окончания должно быть больше времени начала", 400
            except ValueError:
                return "Неверный формат времени окончания", 400
        
        # Проверяем уникальность номера партии
        if Batch.query.filter_by(batch_number=batch_number).first():
            return "Партия с таким номером уже существует", 400
        
        new_batch = Batch(
            user_id=session["user_id"],
            product_id=product_id if product_id else None,
            batch_number=batch_number,
            batch_type=batch_type,
            equipment_id=equipment_id if equipment_id else None,
            notes=notes,
            start_time=start_time,
            end_time=end_time
        )
        
        db.session.add(new_batch)
        db.session.commit()
        clear_analytics_cache()  # Очищаем кэш при создании новой партии
        
        return redirect(url_for("batch_list"))
    
    # Получаем оборудование и продукты для формы
    equipment = Equipment.query.filter_by(status="operational").all()
    products = Product.query.filter_by(is_active=True).all()
    return render_template("create_batch.html", equipment=equipment, products=products)

# === ШАБЛОНЫ ПАРТИЙ ===
@app.route("/templates")
def templates_list():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    templates = BatchTemplate.query.filter_by(is_active=True).all()
    return render_template("templates_list.html", templates=templates)

@app.route("/add_template", methods=["GET", "POST"])
def add_template():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    if request.method == "POST":
        name = request.form["name"]
        batch_type = request.form["batch_type"]
        product_id = request.form.get("product_id", type=int)
        equipment_id = request.form.get("equipment_id", type=int)
        notes = request.form.get("notes", "")
        
        # Создаем шаблон
        template = BatchTemplate(
            name=name,
            batch_type=batch_type,
            product_id=product_id if product_id else None,
            equipment_id=equipment_id if equipment_id else None,
            notes=notes,
            created_by=session["user_id"]
        )
        
        db.session.add(template)
        db.session.flush()  # Получаем ID шаблона
        
        # Добавляем материалы
        materials = Material.query.filter_by(is_active=True).all()
        for material in materials:
            quantity = request.form.get(f"material_{material.id}", type=float)
            if quantity and quantity > 0:
                template_material = BatchTemplateMaterial(
                    template_id=template.id,
                    material_id=material.id,
                    quantity=quantity
                )
                db.session.add(template_material)
        
        db.session.commit()
        return redirect(url_for("templates_list"))
    
    # Получаем данные для формы
    equipment = Equipment.query.filter_by(status="operational").all()
    products = Product.query.filter_by(is_active=True).all()
    materials = Material.query.filter_by(is_active=True).all()
    
    return render_template("add_template.html", equipment=equipment, products=products, materials=materials)

@app.route("/create_batch_from_template/<int:template_id>", methods=["GET", "POST"])
def create_batch_from_template(template_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    template = BatchTemplate.query.get_or_404(template_id)
    
    if request.method == "POST":
        batch_number = request.form["batch_number"]
        notes = request.form.get("notes", "")
        
        # Проверяем уникальность номера партии
        if Batch.query.filter_by(batch_number=batch_number).first():
            return "Партия с таким номером уже существует", 400
        
        # Создаем партию из шаблона
        new_batch = Batch(
            user_id=session["user_id"],
            product_id=template.product_id,
            batch_number=batch_number,
            batch_type=template.batch_type,
            equipment_id=template.equipment_id,
            notes=notes or template.notes,
            start_time=datetime.now()
        )
        
        db.session.add(new_batch)
        db.session.flush()  # Получаем ID партии
        
        # Добавляем материалы из шаблона
        for template_material in template.materials:
            batch_material = BatchMaterial(
                batch_id=new_batch.id,
                material_id=template_material.material_id,
                quantity=template_material.quantity
            )
            db.session.add(batch_material)
        
        db.session.commit()
        return redirect(url_for("batch_list"))
    
    return render_template("create_batch_from_template.html", template=template)

@app.route("/duplicate_last_batch")
def duplicate_last_batch():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    # Находим последнюю партию пользователя
    last_batch = Batch.query.filter_by(user_id=session["user_id"]).order_by(Batch.start_time.desc()).first()
    
    if not last_batch:
        return "Нет партий для дублирования", 400
    
    # Создаем новую партию на основе последней
    new_batch_number = f"{last_batch.batch_number}_copy_{datetime.now().strftime('%H%M%S')}"
    
    new_batch = Batch(
        user_id=session["user_id"],
        product_id=last_batch.product_id,
        batch_number=new_batch_number,
        batch_type=last_batch.batch_type,
        equipment_id=last_batch.equipment_id,
        notes=f"Создано на основе партии {last_batch.batch_number}",
        start_time=datetime.now()
    )
    
    db.session.add(new_batch)
    db.session.flush()
    
    # Копируем материалы
    for material in last_batch.materials:
        new_material = BatchMaterial(
            batch_id=new_batch.id,
            material_id=material.material_id,
            quantity=material.quantity
        )
        db.session.add(new_material)
    
    db.session.commit()
    return redirect(url_for("batch_list"))

# === РЕДАКТИРОВАНИЕ ЗАПИСЕЙ ===
@app.route("/edit_entry/<int:entry_id>", methods=["GET", "POST"])
def edit_entry(entry_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    entry = Entry.query.get_or_404(entry_id)
    
    # Проверяем права доступа
    if entry.user_id != session["user_id"] and session.get("role") not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    # Проверяем время редактирования (30 минут)
    edit_time_limit = 30  # минут
    time_since_creation = (datetime.now() - entry.created_at).total_seconds() / 60
    
    if time_since_creation > edit_time_limit:
        return f"Время редактирования истекло. Можно редактировать только в течение {edit_time_limit} минут после создания.", 400
    
    if request.method == "POST":
        # Обновляем данные
        entry.cement = float(request.form.get("cement", 0))
        entry.lime = float(request.form.get("lime", 0))
        entry.alum_powder = float(request.form.get("alum_powder", 0))
        entry.sludge = float(request.form.get("sludge", 0))
        entry.gypsum = float(request.form.get("gypsum", 0))
        entry.water = float(request.form.get("water", 0))
        entry.sulfanol = float(request.form.get("sulfanol", 0))
        entry.time = request.form.get("time", entry.time)
        entry.shift = request.form.get("shift", entry.shift)
        entry.updated_at = datetime.now()
        
        db.session.commit()
        return redirect(url_for("employee_dashboard"))
    
    return render_template("edit_entry.html", entry=entry, time_limit=edit_time_limit, time_remaining=edit_time_limit - time_since_creation)

@app.route("/edit_batch/<int:batch_id>", methods=["GET", "POST"])
def edit_batch(batch_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем права доступа
    if batch.user_id != session["user_id"] and session.get("role") not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    # Проверяем время редактирования (30 минут)
    edit_time_limit = 30  # минут
    time_since_creation = (datetime.now() - batch.created_at).total_seconds() / 60
    
    if time_since_creation > edit_time_limit:
        return f"Время редактирования истекло. Можно редактировать только в течение {edit_time_limit} минут после создания.", 400
    
    if request.method == "POST":
        # Обновляем данные
        batch.batch_number = request.form.get("batch_number", batch.batch_number)
        batch.notes = request.form.get("notes", batch.notes)
        batch.updated_at = datetime.now()
        
        # Обновляем материалы
        for material in batch.materials:
            quantity = request.form.get(f"material_{material.material_id}", type=float)
            if quantity is not None:
                material.quantity = quantity
        
        db.session.commit()
        return redirect(url_for("batch_list"))
    
    # Получаем данные для формы
    equipment = Equipment.query.filter_by(status="operational").all()
    products = Product.query.filter_by(is_active=True).all()
    
    return render_template("edit_batch.html", batch=batch, equipment=equipment, products=products, 
                          time_limit=edit_time_limit, time_remaining=edit_time_limit - time_since_creation)

@app.route("/batch_list")
def batch_list():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    
    # Получаем параметры фильтрации
    page = request.args.get('page', 1, type=int)
    per_page = 50  # По 50 записей на страницу
    
    # Фильтры
    batch_type_filter = request.args.get('batch_type', 'all')
    status_filter = request.args.get('status', 'all')
    product_filter = request.args.get('product', 'all')
    equipment_filter = request.args.get('equipment', 'all')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Базовый запрос
    if user_role == "admin":
        query = db.session.query(Batch, User).join(User, Batch.user_id == User.id)
    elif user_role in ["director", "chief_technologist"]:
        query = db.session.query(Batch, User).join(User, Batch.user_id == User.id)
    else:
        query = db.session.query(Batch, User).join(User, Batch.user_id == User.id).filter(Batch.user_id == session["user_id"])
    
    # Применяем фильтры
    if batch_type_filter != 'all':
        query = query.filter(Batch.batch_type == batch_type_filter)
    
    if status_filter != 'all':
        query = query.filter(Batch.status == status_filter)
    
    if product_filter != 'all':
        query = query.filter(Batch.product_id == product_filter)
    
    if equipment_filter != 'all':
        query = query.filter(Batch.equipment_id == equipment_filter)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Batch.start_time >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            # Добавляем 23:59:59 к дате "до"
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(Batch.start_time <= to_date)
        except ValueError:
            pass
    
    # Сортировка по времени создания (новые сначала)
    query = query.order_by(Batch.start_time.desc())
    
    # Пагинация
    batches_pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Получаем справочники для фильтров
    products = Product.query.filter_by(is_active=True).all()
    equipment = Equipment.query.all()
    
    return render_template("batch_list.html", 
                         batches=batches_pagination.items,
                         pagination=batches_pagination,
                         products=products,
                         equipment=equipment,
                         current_filters={
                             'batch_type': batch_type_filter,
                             'status': status_filter,
                             'product': product_filter,
                             'equipment': equipment_filter,
                             'date_from': date_from,
                             'date_to': date_to
                         })

@app.route("/batch/<int:batch_id>")
def batch_detail(batch_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    batch = db.session.query(Batch, User).join(User, Batch.user_id == User.id).filter(Batch.id == batch_id).first()
    if not batch:
        return "Партия не найдена", 404
    
    batch_obj, user = batch
    
    # Проверяем права доступа
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"] and batch_obj.user_id != session["user_id"]:
        return "Доступ запрещен", 403
    
    # Получаем материалы партии
    materials = db.session.query(BatchMaterial, Material).join(Material, BatchMaterial.material_id == Material.id).filter(BatchMaterial.batch_id == batch_id).all()
    
    return render_template("batch_detail.html", batch=batch_obj, user=user, materials=materials)

@app.route("/add_material_to_batch/<int:batch_id>", methods=["GET", "POST"])
def add_material_to_batch(batch_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем права доступа
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"] and batch.user_id != session["user_id"]:
        return "Доступ запрещен", 403
    
    if request.method == "POST":
        material_id = request.form["material_id"]
        quantity = request.form.get("quantity", type=float)
        
        batch_material = BatchMaterial(
            batch_id=batch_id,
            material_id=material_id,
            quantity=quantity
        )
        
        db.session.add(batch_material)
        db.session.commit()
        
        return redirect(url_for("batch_detail", batch_id=batch_id))
    
    # Получаем все материалы
    materials = Material.query.filter_by(is_active=True).all()
    return render_template("add_material_to_batch.html", batch=batch, materials=materials)

@app.route("/complete_batch/<int:batch_id>", methods=["POST"])
def complete_batch(batch_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем права доступа
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"] and batch.user_id != session["user_id"]:
        return "Доступ запрещен", 403
    
    batch.status = "completed"
    batch.end_time = datetime.now()
    
    db.session.commit()
    
    return redirect(url_for("batch_detail", batch_id=batch_id))

@app.route("/cancel_batch/<int:batch_id>", methods=["GET", "POST"])
def cancel_batch(batch_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем права доступа
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        # Операторы могут отменять только свои партии
        if batch.user_id != session["user_id"]:
            return "Доступ запрещен", 403
    
    if request.method == "POST":
        reason = request.form.get("reason", "Отменено оператором")
        
        if batch.status == "active":
            batch.status = "cancelled"
            batch.end_time = datetime.now()
            batch.notes = f"{batch.notes}\nОтменено: {reason}" if batch.notes else f"Отменено: {reason}"
            
            db.session.commit()
            
            return redirect(url_for("batch_detail", batch_id=batch_id))
    
    # GET запрос - показываем форму отмены
    return render_template("cancel_batch.html", batch=batch)

@app.route("/pause_batch/<int:batch_id>", methods=["POST"])
def pause_batch(batch_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем права доступа
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        # Операторы могут приостанавливать только свои партии
        if batch.user_id != session["user_id"]:
            return "Доступ запрещен", 403
    
    if batch.status == "active":
        batch.status = "inactive"
        batch.notes = f"{batch.notes}\nПриостановлено: {datetime.now().strftime('%d.%m.%Y %H:%M')}" if batch.notes else f"Приостановлено: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        db.session.commit()
    
    return redirect(url_for("batch_detail", batch_id=batch_id))

@app.route("/resume_batch/<int:batch_id>", methods=["POST"])
def resume_batch(batch_id):
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем права доступа
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        # Операторы могут возобновлять только свои партии
        if batch.user_id != session["user_id"]:
            return "Доступ запрещен", 403
    
    if batch.status == "inactive":
        batch.status = "active"
        batch.notes = f"{batch.notes}\nВозобновлено: {datetime.now().strftime('%d.%m.%Y %H:%M')}" if batch.notes else f"Возобновлено: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        
        db.session.commit()
    
    return redirect(url_for("batch_detail", batch_id=batch_id))

# Специальные маршруты для резчика

@app.route("/start_cutting_batch", methods=["POST"])
def start_cutting_batch():
    if "user_id" not in session or session.get("role") != "cutting_operator":
        return redirect(url_for("login"))
    
    batch_number = request.form["batch_number"]
    equipment_id = request.form.get("equipment_id", type=int)
    shift = request.form.get("shift", "day")
    notes = request.form.get("notes", "")
    
    # Проверяем уникальность номера партии
    if Batch.query.filter_by(batch_number=batch_number).first():
        return "Партия с таким номером уже существует", 400
    
    new_batch = Batch(
        user_id=session["user_id"],
        batch_number=batch_number,
        batch_type="cutting",
        equipment_id=equipment_id if equipment_id else None,
        shift=shift,
        notes=notes,
        status="active"
    )
    
    db.session.add(new_batch)
    db.session.commit()
    
    return redirect(url_for("cutting_dashboard"))

@app.route("/stop_cutting_batch/<int:batch_id>", methods=["POST"])
def stop_cutting_batch(batch_id):
    if "user_id" not in session or session.get("role") != "cutting_operator":
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем, что это партия текущего пользователя
    if batch.user_id != session["user_id"]:
        return "Доступ запрещен", 403
    
    # Проверяем, что это партия резки
    if batch.batch_type != "cutting":
        return "Это не партия резки", 400
    
    batch.status = "completed"
    batch.end_time = datetime.now()
    
    db.session.commit()
    
    return redirect(url_for("cutting_dashboard"))

# Специальные маршруты для автоклавщика

@app.route("/start_autoclave_cycle", methods=["POST"])
def start_autoclave_cycle():
    if "user_id" not in session or session.get("role") != "autoclave_operator":
        return redirect(url_for("login"))
    
    batch_number = request.form["batch_number"]
    equipment_id = request.form.get("equipment_id", type=int)
    shift = request.form.get("shift", "day")
    notes = request.form.get("notes", "")
    
    # Проверяем уникальность номера партии
    if Batch.query.filter_by(batch_number=batch_number).first():
        return "Партия с таким номером уже существует", 400
    
    new_batch = Batch(
        user_id=session["user_id"],
        batch_number=batch_number,
        batch_type="autoclave",
        equipment_id=equipment_id if equipment_id else None,
        shift=shift,
        notes=notes,
        status="active"
    )
    
    db.session.add(new_batch)
    db.session.commit()
    
    return redirect(url_for("autoclave_dashboard"))

@app.route("/stop_autoclave_cycle/<int:batch_id>", methods=["POST"])
def stop_autoclave_cycle(batch_id):
    if "user_id" not in session or session.get("role") != "autoclave_operator":
        return redirect(url_for("login"))
    
    batch = Batch.query.get_or_404(batch_id)
    
    # Проверяем, что это партия текущего пользователя
    if batch.user_id != session["user_id"]:
        return "Доступ запрещен", 403
    
    # Проверяем, что это партия автоклава
    if batch.batch_type != "autoclave":
        return "Это не партия автоклава", 400
    
    batch.status = "completed"
    batch.end_time = datetime.now()
    
    db.session.commit()
    
    return redirect(url_for("autoclave_dashboard"))

# Маршрут для инициализации справочников
@app.route("/init_references")
def init_references():
    if session.get("role") != "admin":
        return "Доступ запрещен", 403
    
    # Создаем материалы
    materials_data = [
        ("Цемент", "kg"),
        ("Известь", "kg"),
        ("Алюминиевая пудра", "kg"),
        ("Шлам", "l"),
        ("Гипс", "kg"),
        ("Вода", "l"),
        ("Сульфанол", "l")
    ]
    
    for name, unit in materials_data:
        if not Material.query.filter_by(name=name).first():
            material = Material(name=name, unit=unit)
            db.session.add(material)
    
    # Создаем оборудование
    equipment_data = [
        ("Смеситель №1", "mixer"),
        ("Смеситель №2", "mixer"),
        ("Резчик №1", "cutter"),
        ("Резчик №2", "cutter"),
        ("Автоклав №1", "autoclave"),
        ("Автоклав №2", "autoclave")
    ]
    
    for name, eq_type in equipment_data:
        if not Equipment.query.filter_by(name=name).first():
            equipment = Equipment(name=name, equipment_type=eq_type)
            db.session.add(equipment)
    
    # Создаем продукты
    products_data = [
        ("BLK-001", "Блок газобетонный D400 600x200x300", "Блоки газобетонные марки D400 размером 600x200x300 мм"),
        ("BLK-002", "Блок газобетонный D500 600x200x300", "Блоки газобетонные марки D500 размером 600x200x300 мм"),
        ("BLK-003", "Блок газобетонный D600 600x200x300", "Блоки газобетонные марки D600 размером 600x200x300 мм"),
        ("PAN-001", "Плита газобетонная D400 600x300x100", "Плиты газобетонные марки D400 размером 600x300x100 мм"),
        ("PAN-002", "Плита газобетонная D500 600x300x150", "Плиты газобетонные марки D500 размером 600x300x150 мм"),
        ("U-BLK-001", "U-блок газобетонный D500 600x200x300", "U-образные блоки газобетонные марки D500 размером 600x200x300 мм"),
        ("U-BLK-002", "U-блок газобетонный D600 600x200x300", "U-образные блоки газобетонные марки D600 размером 600x200x300 мм"),
        ("WALL-001", "Стеновой блок D500 600x200x400", "Стеновые блоки газобетонные марки D500 размером 600x200x400 мм")
    ]
    
    for code, name, description in products_data:
        if not Product.query.filter_by(product_code=code).first():
            product = Product(product_code=code, name=name, description=description)
            db.session.add(product)
    
    db.session.commit()
    return "Справочники инициализированы (материалы, оборудование, продукты)"

# === ЭКСПОРТ ДАННЫХ ===
@app.route("/export/batches_csv")
def export_batches_csv():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    # Получаем параметры фильтрации
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    batch_type_filter = request.args.get('batch_type', 'all')
    status_filter = request.args.get('status', 'all')
    product_filter = request.args.get('product', 'all')
    equipment_filter = request.args.get('equipment', 'all')
    
    # Строим запрос
    query = db.session.query(Batch, User).join(User, Batch.user_id == User.id)
    
    if batch_type_filter != 'all':
        query = query.filter(Batch.batch_type == batch_type_filter)
    
    if status_filter != 'all':
        query = query.filter(Batch.status == status_filter)
    
    if product_filter != 'all':
        query = query.filter(Batch.product_id == product_filter)
    
    if equipment_filter != 'all':
        query = query.filter(Batch.equipment_id == equipment_filter)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(Batch.start_time >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(Batch.start_time <= to_date)
        except ValueError:
            pass
    
    batches = query.order_by(Batch.start_time.desc()).all()
    
    # Создаем Excel файл с форматированием
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Производственные партии"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = [
        'ID', 'Номер партии', 'Тип', 'Статус', 'Продукт', 'Оборудование',
        'Оператор', 'Время начала', 'Время окончания', 'Примечания'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Данные
    for row, (batch, user) in enumerate(batches, 2):
        data = [
            batch.id,
            batch.batch_number,
            batch.batch_type,
            batch.status,
            batch.product.name if batch.product else '',
            batch.equipment.name if batch.equipment else '',
            user.fio,
            batch.start_time.strftime('%d.%m.%Y %H:%M'),
            batch.end_time.strftime('%d.%m.%Y %H:%M') if batch.end_time else '',
            batch.notes or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Автоширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Замораживаем первую строку
    worksheet.freeze_panes = "A2"
    
    workbook.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=batches_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@app.route("/export/entries_csv")
def export_entries_csv():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    # Получаем параметры фильтрации
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    shift_filter = request.args.get('shift', 'all')
    
    # Строим запрос
    query = db.session.query(Entry, User).join(User, Entry.user_id == User.id)
    
    if shift_filter != 'all':
        query = query.filter(Entry.shift == shift_filter)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Entry.date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Entry.date <= to_date)
        except ValueError:
            pass
    
    entries = query.order_by(Entry.date.desc(), Entry.time.desc()).all()
    
    # Создаем Excel файл с форматированием
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Записи ввода материалов"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки
    headers = [
        'ID', 'Дата', 'Время', 'Смена', 'Оператор',
        'Цемент (кг)', 'Известь (кг)', 'Ал. пудра (кг)', 'Шлам (л)',
        'Гипс (кг)', 'Вода (л)', 'Сульфанол (л)'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Данные
    for row, (entry, user) in enumerate(entries, 2):
        data = [
            entry.id,
            entry.date.strftime('%d.%m.%Y'),
            entry.time,
            entry.shift,
            user.fio,
            entry.cement,
            entry.lime,
            entry.alum_powder,
            entry.sludge,
            entry.gypsum,
            entry.water,
            entry.sulfanol
        ]
        
        for col, value in enumerate(data, 1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Форматирование числовых значений
            if col >= 6:  # Материалы
                if value and str(value).replace('.', '').replace(',', '').isdigit():
                    cell.number_format = '0.00'
    
    # Автоширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Замораживаем первую строку
    worksheet.freeze_panes = "A2"
    
    workbook.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=entries_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

@app.route("/export/analytics_csv")
def export_analytics_csv():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    # Получаем параметры фильтрации
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # Строим запрос для партий
    batch_query = db.session.query(Batch, User).join(User, Batch.user_id == User.id)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            batch_query = batch_query.filter(Batch.start_time >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            batch_query = batch_query.filter(Batch.start_time <= to_date)
        except ValueError:
            pass
    
    batches = batch_query.all()
    
    # Создаем Excel файл с аналитикой
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Аналитика производства"
    
    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    summary_font = Font(bold=True, color="FFFFFF", size=11)
    summary_fill = PatternFill(start_color="5D4037", end_color="5D4037", fill_type="solid")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовки аналитики по типам
    headers = ['Тип партии', 'Количество', 'Завершено', 'Активно', 'Отменено', 'Приостановлено']
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Аналитика по типам партий
    batch_types = ['casting', 'cutting', 'autoclave']
    type_names = {'casting': 'Заливка', 'cutting': 'Резка', 'autoclave': 'Автоклав'}
    
    for row, batch_type in enumerate(batch_types, 2):
        type_batches = [b for b, u in batches if b.batch_type == batch_type]
        total = len(type_batches)
        completed = len([b for b in type_batches if b.status == 'completed'])
        active = len([b for b in type_batches if b.status == 'active'])
        cancelled = len([b for b in type_batches if b.status == 'cancelled'])
        inactive = len([b for b in type_batches if b.status == 'inactive'])
        
        data = [type_names.get(batch_type, batch_type), total, completed, active, cancelled, inactive]
        
        for col, value in enumerate(data, 1):
            cell = worksheet.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Общая статистика
    summary_row = len(batch_types) + 3
    worksheet.cell(row=summary_row, column=1, value="ОБЩАЯ СТАТИСТИКА").font = summary_font
    worksheet.cell(row=summary_row, column=1).fill = summary_fill
    worksheet.cell(row=summary_row, column=1).border = thin_border
    
    summary_data = [
        ['Всего партий', len(batches)],
        ['Завершено', len([b for b, u in batches if b.status == 'completed'])],
        ['Активно', len([b for b, u in batches if b.status == 'active'])],
        ['Отменено', len([b for b, u in batches if b.status == 'cancelled'])],
        ['Приостановлено', len([b for b, u in batches if b.status == 'inactive'])]
    ]
    
    for i, (label, value) in enumerate(summary_data, 1):
        row = summary_row + i
        worksheet.cell(row=row, column=1, value=label).border = thin_border
        worksheet.cell(row=row, column=2, value=value).border = thin_border
        worksheet.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
    
    # Автоширина колонок
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Замораживаем первую строку
    worksheet.freeze_panes = "A2"
    
    workbook.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=analytics_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

# === УПРАВЛЕНИЕ КЭШЕМ ===
@app.route("/cache/clear")
def clear_cache():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    cache.clear()
    return "Кэш очищен", 200

@app.route("/cache/status")
def cache_status():
    if "user_id" not in session:
        return redirect(url_for("login"))
    
    user_role = session.get("role")
    if user_role not in ["admin", "director", "chief_technologist"]:
        return "Доступ запрещен", 403
    
    status = {
        "cache_size": len(cache.cache),
        "cached_keys": list(cache.cache.keys()),
        "cache_ttl": {k: cache.cache_ttl.get(k, 0) - time.time() for k in cache.cache.keys()}
    }
    return jsonify(status)

# === ОПТИМИЗИРОВАННЫЕ ФУНКЦИИ ДЛЯ КЭШИРОВАНИЯ ===
def get_daily_analytics_data():
    """Получает данные для ежедневной аналитики"""
    today = datetime.now().date()
    
    # Записи за сегодня
    today_entries = db.session.query(Entry, User).join(User, Entry.user_id == User.id).filter(
        Entry.date == today
    ).all()
    
    # Партии за сегодня
    today_batches = db.session.query(Batch, User).join(User, Batch.user_id == User.id).filter(
        db.func.date(Batch.start_time) == today
    ).all()
    
    return {
        "today_entries": today_entries,
        "today_batches": today_batches,
        "total_entries": len(today_entries),
        "total_batches": len(today_batches)
    }

def get_weekly_analytics_data():
    """Получает данные для недельной аналитики"""
    week_ago = datetime.now() - timedelta(days=7)
    
    # Записи за неделю
    week_entries = db.session.query(Entry, User).join(User, Entry.user_id == User.id).filter(
        Entry.date >= week_ago.date()
    ).all()
    
    # Партии за неделю
    week_batches = db.session.query(Batch, User).join(User, Batch.user_id == User.id).filter(
        Batch.start_time >= week_ago
    ).all()
    
    return {
        "week_entries": week_entries,
        "week_batches": week_batches,
        "total_entries": len(week_entries),
        "total_batches": len(week_batches)
    }

def get_material_consumption_data():
    """Получает данные о расходе материалов"""
    # Материалы из Entry записей
    entry_materials = db.session.query(
        db.literal('entry').label('source'),
        db.literal('casting').label('batch_type'),
        db.literal('Цемент').label('material_name'),
        db.literal('kg').label('unit'),
        db.func.sum(Entry.cement).label('total_quantity')
    ).group_by(db.literal(1)).union_all(
        db.session.query(
            db.literal('entry').label('source'),
            db.literal('casting').label('batch_type'),
            db.literal('Известь').label('material_name'),
            db.literal('kg').label('unit'),
            db.func.sum(Entry.lime).label('total_quantity')
        ).group_by(db.literal(1))
    ).union_all(
        db.session.query(
            db.literal('entry').label('source'),
            db.literal('casting').label('batch_type'),
            db.literal('Алюминиевая пудра').label('material_name'),
            db.literal('kg').label('unit'),
            db.func.sum(Entry.alum_powder).label('total_quantity')
        ).group_by(db.literal(1))
    ).union_all(
        db.session.query(
            db.literal('entry').label('source'),
            db.literal('casting').label('batch_type'),
            db.literal('Шлам').label('material_name'),
            db.literal('l').label('unit'),
            db.func.sum(Entry.sludge).label('total_quantity')
        ).group_by(db.literal(1))
    ).union_all(
        db.session.query(
            db.literal('entry').label('source'),
            db.literal('casting').label('batch_type'),
            db.literal('Гипс').label('material_name'),
            db.literal('kg').label('unit'),
            db.func.sum(Entry.gypsum).label('total_quantity')
        ).group_by(db.literal(1))
    ).union_all(
        db.session.query(
            db.literal('entry').label('source'),
            db.literal('casting').label('batch_type'),
            db.literal('Вода').label('material_name'),
            db.literal('l').label('unit'),
            db.func.sum(Entry.water).label('total_quantity')
        ).group_by(db.literal(1))
    ).union_all(
        db.session.query(
            db.literal('entry').label('source'),
            db.literal('casting').label('batch_type'),
            db.literal('Сульфанол').label('material_name'),
            db.literal('l').label('unit'),
            db.func.sum(Entry.sulfanol).label('total_quantity')
        ).group_by(db.literal(1))
    )

    # Материалы из BatchMaterial
    batch_materials = db.session.query(
        db.literal('batch').label('source'),
        Batch.batch_type,
        Material.name.label('material_name'),
        Material.unit,
        db.func.sum(BatchMaterial.quantity).label('total_quantity')
    ).join(
        BatchMaterial, Batch.id == BatchMaterial.batch_id
    ).join(
        Material, BatchMaterial.material_id == Material.id
    ).group_by(
        Batch.batch_type,
        Material.name,
        Material.unit
    )

    # Объединяем все источники
    all_materials_query = entry_materials.union_all(batch_materials)
    all_materials_data = all_materials_query.all()

    # Структурируем данные
    materials_by_batch_type = {
        'casting': {},
        'cutting': {},
        'autoclave': {}
    }

    materials_total = {}

    for source, batch_type, material_name, unit, total_qty in all_materials_data:
        if not material_name or not total_qty:
            continue

        # Добавляем в общий словарь
        if material_name not in materials_total:
            materials_total[material_name] = {
                'unit': unit,
                'quantity': 0
            }
        materials_total[material_name]['quantity'] += float(total_qty) if total_qty else 0

        # Добавляем в словарь по типам партий
        if batch_type and batch_type in materials_by_batch_type:
            if material_name not in materials_by_batch_type[batch_type]:
                materials_by_batch_type[batch_type][material_name] = {
                    'unit': unit,
                    'quantity': 0
                }
            materials_by_batch_type[batch_type][material_name]['quantity'] += float(total_qty) if total_qty else 0

    return {
        "materials_by_batch_type": materials_by_batch_type,
        "materials_total": materials_total
    }

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    
    # Для Railway и продакшена
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(host='0.0.0.0', port=port, debug=debug)